from __future__ import annotations

import re
from pathlib import Path

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.utils.storage import new_file_id, output_path


class BankError(Exception):
    pass


def _normalize(rows: list[list]) -> list[list[str]]:
    cleaned: list[list[str]] = []
    for row in rows:
        if not row:
            continue
        normalized = [(c or "").strip() if isinstance(c, str) else ("" if c is None else str(c)) for c in row]
        if any(normalized):
            cleaned.append(normalized)
    return cleaned


def _detect_header(rows: list[list[str]]) -> int:
    """Return index of the header row (best heuristic)."""
    keywords = (
        "date", "transaction", "particular", "description", "narration",
        "withdrawal", "deposit", "debit", "credit", "balance", "amount", "reference",
    )
    for i, row in enumerate(rows[:8]):
        joined = " | ".join(c.lower() for c in row)
        if sum(1 for k in keywords if k in joined) >= 2:
            return i
    return -1


def _is_summary_row(row: list[str]) -> bool:
    joined = " ".join(row).lower()
    return any(
        s in joined
        for s in ("opening balance", "closing balance", "total", "carried forward", "brought forward")
    )


def convert_bank_statement(file_path: Path, output_name: str = "statement.xlsx") -> tuple[str, Path, dict]:
    """Extract bank statement tables to a clean XLSX file."""
    try:
        with pdfplumber.open(file_path) as pdf:
            all_rows: list[list[str]] = []
            header: list[str] | None = None
            for page in pdf.pages:
                tables = page.extract_tables() or []
                for tbl in tables:
                    rows = _normalize(tbl)
                    if not rows:
                        continue
                    h_idx = _detect_header(rows)
                    if h_idx >= 0 and header is None:
                        header = rows[h_idx]
                        all_rows.extend(rows[h_idx + 1 :])
                    elif h_idx >= 0:
                        all_rows.extend(rows[h_idx + 1 :])
                    else:
                        all_rows.extend(rows)
            if not all_rows and not header:
                # fallback: try page text → lines with numbers
                lines: list[list[str]] = []
                for page in pdf.pages:
                    text = page.extract_text() or ""
                    for line in text.splitlines():
                        if re.search(r"\d", line):
                            parts = re.split(r"\s{2,}|\t", line.strip())
                            if len(parts) >= 3:
                                lines.append(parts)
                if not lines:
                    raise BankError(
                        "Could not detect a table in this PDF. Ensure the statement has selectable text, image-only scans need OCR first."
                    )
                header = [f"Column {i+1}" for i in range(max(len(r) for r in lines))]
                all_rows = lines
    except BankError:
        raise
    except Exception as exc:  # noqa: BLE001
        raise BankError(f"Failed to read PDF: {exc}") from exc

    # Pad rows to header width and drop summary lines
    max_cols = max((len(r) for r in all_rows), default=0)
    if header:
        max_cols = max(max_cols, len(header))
    cleaned: list[list[str]] = []
    for row in all_rows:
        if _is_summary_row(row):
            continue
        padded = row + [""] * (max_cols - len(row))
        cleaned.append(padded[:max_cols])

    if header is None:
        header = [f"Column {i+1}" for i in range(max_cols)]
    else:
        header = header + [""] * (max_cols - len(header))
        header = header[:max_cols]

    wb = Workbook()
    ws = wb.active
    ws.title = "Statement"

    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill("solid", fgColor="FF4F46E5")
    align_center = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.append(header)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center

    for row in cleaned:
        ws.append(row)

    # Auto width
    for col in range(1, max_cols + 1):
        letter = get_column_letter(col)
        max_len = max(
            (len(str(ws.cell(r, col).value or "")) for r in range(1, ws.max_row + 1)),
            default=10,
        )
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 48)

    ws.freeze_panes = "A2"

    output_id = new_file_id()
    out = output_path(output_id).with_suffix(".xlsx")
    wb.save(out)

    stats = {"rows": len(cleaned), "columns": max_cols, "filename": output_name}
    return output_id, out, stats
