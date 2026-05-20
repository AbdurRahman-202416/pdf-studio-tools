from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape, portrait
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from app.utils.storage import new_file_id, output_path


class ExcelToPDFError(Exception):
    pass


HEADER_BG = colors.HexColor("#1F2937")
HEADER_FG = colors.white
GRID = colors.HexColor("#E2E8F0")
ZEBRA = colors.HexColor("#F8FAFC")

WIDE_THRESHOLD_COLUMNS = 6


def _sheet_rows(ws) -> list[list[str]]:
    rows: list[list[str]] = []
    for r in ws.iter_rows(values_only=True):
        cells = ["" if c is None else str(c) for c in r]
        if any(c.strip() for c in cells):
            rows.append(cells)
    return rows


def _decide_orientation(rows: list[list[str]]) -> bool:
    if not rows:
        return False
    max_cols = max(len(r) for r in rows)
    if max_cols > WIDE_THRESHOLD_COLUMNS:
        return True
    avg_chars = sum(sum(len(c) for c in row) for row in rows) / max(len(rows), 1)
    return avg_chars > 80


def _build_table(rows: list[list[str]], available_width: float) -> Table:
    max_cols = max(len(r) for r in rows)
    padded = [r + [""] * (max_cols - len(r)) for r in rows]

    body_style = ParagraphStyle(
        "cell",
        fontName="Helvetica",
        fontSize=8.5,
        leading=10.5,
        spaceBefore=0,
        spaceAfter=0,
    )
    header_style = ParagraphStyle(
        "cellHeader",
        fontName="Helvetica-Bold",
        fontSize=8.5,
        leading=10.5,
        textColor=HEADER_FG,
    )

    wrapped: list[list] = []
    for row_idx, row in enumerate(padded):
        wrapped_row = []
        for cell in row:
            text = (cell or "").replace("\n", "<br/>")
            style = header_style if row_idx == 0 else body_style
            wrapped_row.append(Paragraph(text or "&nbsp;", style))
        wrapped.append(wrapped_row)

    col_widths = [available_width / max_cols] * max_cols
    table = Table(wrapped, colWidths=col_widths, repeatRows=1)

    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), HEADER_BG),
        ("TEXTCOLOR", (0, 0), (-1, 0), HEADER_FG),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("GRID", (0, 0), (-1, -1), 0.4, GRID),
    ]
    for i in range(1, len(padded)):
        if i % 2 == 0:
            style_cmds.append(("BACKGROUND", (0, i), (-1, i), ZEBRA))
    table.setStyle(TableStyle(style_cmds))
    return table


def convert_excel_to_pdf(file_path: Path, output_name: str = "spreadsheet.pdf") -> tuple[str, Path, dict]:
    """Render every sheet in the XLSX as one PDF, one sheet per page (or more if it overflows).

    Behavior:
      - Each sheet starts on a new page.
      - Sheet title rendered as a heading.
      - Tables get a styled header, zebra rows, borders, and column repeat across pages.
      - Wide sheets are rendered in landscape.
    """
    try:
        wb = load_workbook(file_path, data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001
        raise ExcelToPDFError(f"Failed to open spreadsheet: {exc}") from exc

    if not wb.sheetnames:
        raise ExcelToPDFError("This workbook has no sheets")

    sheets: list[tuple[str, list[list[str]]]] = []
    any_landscape = False
    for name in wb.sheetnames:
        rows = _sheet_rows(wb[name])
        if not rows:
            continue
        sheets.append((name, rows))
        if not any_landscape and _decide_orientation(rows):
            any_landscape = True

    if not sheets:
        raise ExcelToPDFError("All sheets in this workbook are empty")

    pagesize = landscape(A4) if any_landscape else portrait(A4)
    margin = 12 * mm
    available_width = pagesize[0] - 2 * margin

    output_id = new_file_id()
    out = output_path(output_id).with_suffix(".pdf")

    doc = SimpleDocTemplate(
        str(out),
        pagesize=pagesize,
        leftMargin=margin,
        rightMargin=margin,
        topMargin=margin,
        bottomMargin=margin,
        title=output_name,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "sheetTitle",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=13,
        textColor=colors.HexColor("#1F2937"),
        spaceAfter=6,
    )

    story = []
    total_rows = 0
    for idx, (name, rows) in enumerate(sheets):
        if idx > 0:
            story.append(PageBreak())
        story.append(Paragraph(name, title_style))
        story.append(Spacer(1, 4))
        story.append(_build_table(rows, available_width))
        # Header row is excluded from the count to match pdf_table_service.
        total_rows += max(len(rows) - 1, 0)

    doc.build(story)

    stats = {
        "sheets": len(sheets),
        "rows": total_rows,
        "orientation": "landscape" if any_landscape else "portrait",
        "filename": output_name,
    }
    return output_id, out, stats
