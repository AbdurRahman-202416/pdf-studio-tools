from __future__ import annotations

import re
from pathlib import Path

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.utils.guards import assert_pdf_page_limit
from app.utils.storage import new_file_id, output_path


class PDFTableError(Exception):
    pass


HEADER_FILL = PatternFill("solid", fgColor="FF1F2937")
HEADER_FONT = Font(bold=True, color="FFFFFFFF", size=11)
TITLE_FONT = Font(bold=True, color="FF1F2937", size=12)
ZEBRA_FILL = PatternFill("solid", fgColor="FFF8FAFC")
THIN_BORDER = Border(
    left=Side(style="thin", color="FFE2E8F0"),
    right=Side(style="thin", color="FFE2E8F0"),
    top=Side(style="thin", color="FFE2E8F0"),
    bottom=Side(style="thin", color="FFE2E8F0"),
)
LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)
LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)


_HEADER_KEYWORDS = (
    "date", "transaction", "particular", "particulars", "description", "narration",
    "withdrawal", "deposit", "debit", "credit", "balance", "amount", "reference",
    "no.", "qty", "quantity", "rate", "price", "total", "subtotal", "tax", "gst",
    "item", "product", "code", "sku", "id", "name", "category", "type", "status",
)


def _normalize(rows: list[list]) -> list[list[str]]:
    cleaned: list[list[str]] = []
    for row in rows:
        if not row:
            continue
        normalized = [(c or "").strip() if isinstance(c, str) else ("" if c is None else str(c)) for c in row]
        if any(normalized):
            cleaned.append(normalized)
    return cleaned


def _looks_like_header(row: list[str]) -> bool:
    joined = " | ".join(c.lower() for c in row if c)
    if not joined:
        return False
    if sum(1 for k in _HEADER_KEYWORDS if k in joined) >= 2:
        return True
    cells = [c for c in row if c]
    if len(cells) >= 3 and all(len(c) <= 24 and not re.search(r"\d", c) for c in cells):
        return True
    return False


def _split_header(rows: list[list[str]]) -> tuple[list[str] | None, list[list[str]]]:
    for i, row in enumerate(rows[:4]):
        if _looks_like_header(row):
            return row, rows[i + 1 :]
    return None, rows


def _pad(rows: list[list[str]], width: int) -> list[list[str]]:
    return [r + [""] * (width - len(r)) for r in rows]


def _autofit(ws: Worksheet) -> None:
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        widest = 10
        for r in range(1, ws.max_row + 1):
            v = ws.cell(r, col).value
            if v is None:
                continue
            line = max((len(str(s)) for s in str(v).splitlines()), default=0)
            if line > widest:
                widest = line
        ws.column_dimensions[letter].width = min(max(widest + 2, 12), 48)


def _write_table_block(
    ws: Worksheet,
    rows: list[list[str]],
    *,
    start_row: int,
    title: str | None,
) -> int:
    """Render one table block onto a sheet starting at start_row. Returns next free row."""
    if not rows:
        return start_row

    header, body = _split_header(rows)
    width = max((len(r) for r in rows), default=0)
    if header is not None:
        header = (header + [""] * (width - len(header)))[:width]
    else:
        header = [f"Column {i + 1}" for i in range(width)]
    body = _pad(body, width)

    row_cursor = start_row

    if title:
        ws.cell(row_cursor, 1, title).font = TITLE_FONT
        ws.merge_cells(start_row=row_cursor, end_row=row_cursor, start_column=1, end_column=max(width, 1))
        row_cursor += 1

    for col_idx, value in enumerate(header, start=1):
        cell = ws.cell(row_cursor, col_idx, value)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = LEFT_WRAP
        cell.border = THIN_BORDER
    row_cursor += 1

    for i, row in enumerate(body):
        for col_idx, value in enumerate(row, start=1):
            # openpyxl types a leading "=" as a live formula - PDF-extracted
            # text must never execute when the workbook opens.
            if isinstance(value, str) and value.startswith("="):
                value = "'" + value
            cell = ws.cell(row_cursor, col_idx, value)
            cell.alignment = LEFT_TOP
            cell.border = THIN_BORDER
            if i % 2 == 1:
                cell.fill = ZEBRA_FILL
        row_cursor += 1

    return row_cursor + 1


def _extract_pages(pdf: pdfplumber.PDF) -> list[list[list[list[str]]]]:
    """Return [page_idx → list of tables (each table = list of rows)]."""
    pages: list[list[list[list[str]]]] = []
    for page in pdf.pages:
        tables = page.extract_tables() or []
        cleaned_tables = []
        for tbl in tables:
            rows = _normalize(tbl)
            if rows:
                cleaned_tables.append(rows)
        pages.append(cleaned_tables)
    return pages


def _fallback_text_table(pdf: pdfplumber.PDF) -> list[list[list[str]]] | None:
    """Used when pdfplumber finds no tables: split text lines by 2+ spaces or tab to recover columns."""
    lines: list[list[str]] = []
    for page in pdf.pages:
        text = page.extract_text() or ""
        for line in text.splitlines():
            if re.search(r"\d", line):
                parts = re.split(r"\s{2,}|\t", line.strip())
                if len(parts) >= 3:
                    lines.append(parts)
    if not lines:
        return None
    return [lines]


def convert_pdf_table(file_path: Path, output_name: str = "extracted-tables.xlsx") -> tuple[str, Path, dict]:
    """Extract tables from a PDF to a clean, well-formatted XLSX file.

    Behavior:
      - Single-page PDF with N tables → 1 sheet containing all N tables stacked.
      - Multi-page PDF with tables on multiple pages → 1 sheet per page named "Page N".
      - Each table gets a styled header band, zebra rows, borders, autofit columns, frozen header.
    """
    # pdfplumber's table detection is superlinear in vector ops; cap pages first.
    assert_pdf_page_limit(file_path)
    try:
        with pdfplumber.open(file_path) as pdf:
            pages_tables = _extract_pages(pdf)
            total_tables = sum(len(t) for t in pages_tables)
            if total_tables == 0:
                fallback = _fallback_text_table(pdf)
                if not fallback:
                    raise PDFTableError(
                        "Could not detect a table in this PDF. Ensure the PDF has selectable text; "
                        "image-only scans need OCR first."
                    )
                pages_tables = [fallback]
                total_tables = 1
    except PDFTableError:
        raise
    except Exception as exc:  # noqa: BLE001
        raise PDFTableError(f"Failed to read PDF: {exc}") from exc

    pages_with_tables = [i for i, t in enumerate(pages_tables) if t]
    multi_page = len(pages_with_tables) > 1

    if multi_page:
        sheet_plan: list[tuple[str, list[list[list[str]]]]] = [
            (f"Page {i + 1}", pages_tables[i]) for i in pages_with_tables
        ]
    else:
        all_tables = [t for tables in pages_tables for t in tables]
        sheet_plan = [("Tables", all_tables)]

    wb = Workbook()
    wb.remove(wb.active)

    total_rows = 0
    max_cols_seen = 0

    for sheet_title, tables in sheet_plan:
        ws = wb.create_sheet(title=sheet_title)
        row_cursor = 1
        for table_idx, rows in enumerate(tables, start=1):
            block_title = f"Table {table_idx}" if len(tables) > 1 else None
            row_cursor = _write_table_block(ws, rows, start_row=row_cursor, title=block_title)
            total_rows += max(len(rows) - 1, 0)
            max_cols_seen = max(max_cols_seen, max((len(r) for r in rows), default=0))
        ws.freeze_panes = "A3" if len(tables) > 1 else "A2"
        _autofit(ws)

    total_sheets = len(sheet_plan)

    output_id = new_file_id()
    out = output_path(output_id).with_suffix(".xlsx")
    wb.save(out)

    stats = {
        "rows": total_rows,
        "columns": max_cols_seen,
        "sheets": total_sheets,
        "tables": total_tables,
        "filename": output_name,
    }
    return output_id, out, stats
